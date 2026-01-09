"""
Draft Exporter for Jigyokei Electronic Application
Exports the current plan to an Excel sheet with severity-based highlighting.
"""
from io import BytesIO
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from src.api.schemas import ApplicationRoot


class DraftExporter:
    """Export ApplicationRoot to Excel draft sheet"""
    
    @staticmethod
    def export_to_excel(plan: ApplicationRoot, result: Dict[str, Any]) -> BytesIO:
        """
        Create Excel draft sheet with severity-based highlighting.
        
        Args:
            plan: ApplicationRoot instance
            result: Analysis result from CompletionChecker
        
        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "電子申請下書き"
        
        # Define styles
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        critical_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        complete_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Helper: Get severity for a field
        missing_map = {m['section']: m.get('severity', 'complete') for m in result.get('missing_mandatory', [])}
        
        def write_row(row_num, label, value, section_key):
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value if value else "")
            
            # Apply severity-based highlighting
            severity = missing_map.get(section_key, 'complete')
            fill = complete_fill
            if severity == 'critical':
                fill = critical_fill
            elif severity == 'warning':
                fill = warning_fill
            
            for col in [1, 2]:
                cell = ws.cell(row=row_num, column=col)
                cell.fill = fill
                cell.border = border
        
        row = 1
        
        # Header
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws.cell(row=row, column=1, value="事業継続力強化計画 申請下書きシート")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # Section 1: Basic Info
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【様式第1】 基本情報").font = Font(bold=True)
        row += 1
        
        bi = plan.basic_info
        write_row(row, "事業者の氏名又は名称", bi.corporate_name, "BasicInfo"); row += 1
        write_row(row, "事業者の氏名又は名称（フリガナ）", bi.corporate_name_kana, "BasicInfo"); row += 1
        write_row(row, "法人番号", bi.corporate_number, "BasicInfo"); row += 1
        write_row(row, "設立年月日", bi.establishment_date, "BasicInfo"); row += 1
        write_row(row, "代表者の役職", bi.representative_title, "BasicInfo"); row += 1
        write_row(row, "代表者の氏名", f"{bi.representative_name or ''}".strip(), "BasicInfo"); row += 1
        write_row(row, "郵便番号", bi.address_zip, "BasicInfo"); row += 1
        write_row(row, "都道府県", bi.address_pref, "BasicInfo"); row += 1
        write_row(row, "市区町村", bi.address_city, "BasicInfo"); row += 1
        write_row(row, "住所（字・番地等）", bi.address_street, "BasicInfo"); row += 1
        write_row(row, "マンション名等", bi.address_building, "BasicInfo"); row += 1
        write_row(row, "業種（大分類）", bi.industry_major, "BasicInfo"); row += 1
        write_row(row, "業種（中分類）", bi.industry_middle, "BasicInfo"); row += 1
        write_row(row, "資本金又は出資の額", bi.capital, "BasicInfo"); row += 1
        write_row(row, "常時使用する従業員の数", bi.employees, "BasicInfo"); row += 1
        row += 1
        
        # Section 2: Goals
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【様式第2】 事業概要・目標").font = Font(bold=True)
        row += 1
        
        write_row(row, "事業活動の概要", plan.goals.business_overview, "Goals"); row += 1
        write_row(row, "事業継続力強化に取り組む目的", plan.goals.business_purpose, "Goals"); row += 1
        row += 1
        
        # Section 3: Disaster
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【様式第3】 事業活動に影響を与える自然災害等の想定").font = Font(bold=True)
        row += 1
        
        # New Impact Structure
        ds = plan.goals.disaster_scenario
        write_row(row, "事業活動に影響を与える自然災害等の想定", ds.disaster_assumption, "Goals"); row += 1
        
        # Impacts
        imp = ds.impacts
        # Note: Section title inside logic might be implicit, but fields need full name
        write_row(row, "自然災害等の発生が事業活動に与える影響 (人員)", imp.impact_personnel, "Goals"); row += 1
        write_row(row, "自然災害等の発生が事業活動に与える影響 (建物・設備)", imp.impact_building, "Goals"); row += 1
        write_row(row, "自然災害等の発生が事業活動に与える影響 (資金繰り)", imp.impact_funds, "Goals"); row += 1
        write_row(row, "自然災害等の発生が事業活動に与える影響 (情報)", imp.impact_info, "Goals"); row += 1
        row += 1
        
        # Section 4: Response
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【様式第4】 初動対応").font = Font(bold=True)
        row += 1
        
        
        def safe_get(obj, attr, default=""):
            # Helper to get value from object, dict, or tuple (heuristic)
            if hasattr(obj, attr):
                return getattr(obj, attr)
            if isinstance(obj, dict):
                return obj.get(attr, default)
            # Tuple fallback? (Dangerous but maybe debug related)
            return default

        # Strict Order for Response Procedures
        # 1. Safety (Evacuation), 2. Safety (Confirmation), 3. Emergency Org, 4. Damage Info
        
        # Helper to find item by substring of category or content keyword if needed
        # But ideally we rely on the exact category name from AI.
        # Fallback: Search in the list for matching items.
        
        rp_list = plan.response_procedures
        
        def find_rp(keywords):
            for p in rp_list:
                cat = safe_get(p, "category", "")
                cnt = safe_get(p, "action_content", "")
                if any(k in cat for k in keywords):
                    return p
            return None

        # 1. 人命の安全確保 (避難) - Look for "避難" in content or "人命" in category
        # Since we might have 2 "人命" categories, we need to distinguish by content if possible, 
        # OR rely on the order if AI outputs correctly.
        # Strategy: Filter all "人命" keys.
        safety_items = [p for p in rp_list if "人命" in safe_get(p, "category", "") or "Safety" in safe_get(p, "category", "")]
        
        # Evacuation
        evacuation_item = next((p for p in safety_items if "避難" in safe_get(p, "action_content", "")), None)
        if not evacuation_item and safety_items: evacuation_item = safety_items[0] # Fallback
        
        # Confirmation
        confirmation_item = next((p for p in safety_items if "安否" in safe_get(p, "action_content", "")), None)
        if not confirmation_item and len(safety_items) > 1: confirmation_item = safety_items[1] # Fallback
        
        # Emergency Org
        org_item = next((p for p in rp_list if "体制" in safe_get(p, "category", "")), None)
        
        # Damage Info
        info_item = next((p for p in rp_list if "情報" in safe_get(p, "category", "") or "把握" in safe_get(p, "category", "")), None)

        # Write Rows
        # Row 1: Safety (Evacuation)
        write_row(row, "1. 人命の安全確保 (従業員の避難方法)", safe_get(evacuation_item, "action_content"), "ResponseProcedures")
        write_row(row, "   発災後の対応時期", safe_get(evacuation_item, "timing"), "ResponseProcedures")
        row += 1
        
        # Row 2: Safety (Confirmation)
        write_row(row, "   人命の安全確保 (従業員の安否確認)", safe_get(confirmation_item, "action_content"), "ResponseProcedures")
        write_row(row, "   発災後の対応時期", safe_get(confirmation_item, "timing"), "ResponseProcedures")
        row += 1

        # Row 3: Emergency Org
        write_row(row, "2. 非常時の緊急時体制の構築", safe_get(org_item, "action_content"), "ResponseProcedures")
        write_row(row, "   発災後の対応時期", safe_get(org_item, "timing"), "ResponseProcedures")
        row += 1

        # Row 4: Damage Info
        write_row(row, "3. 被害状況の把握・被害情報の共有", safe_get(info_item, "action_content"), "ResponseProcedures")
        write_row(row, "   発災後の対応時期", safe_get(info_item, "timing"), "ResponseProcedures")
        row += 1

        # Cooperation Partners (Separated)
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="協力者").font = Font(bold=True)
        row += 1
        
        for i, cp in enumerate(plan.cooperation_partners, 1):
            name = safe_get(cp, "name")
            content = safe_get(cp, "content")
            write_row(row, f"{i}. 名称", name, "Cooperation"); row += 1
            write_row(row, f"   協力の内容", content, "Cooperation"); row += 1
        row += 1

        
        # Section 5: Measures (A/B/C/D)
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【様式第5】 事前対策").font = Font(bold=True)
        row += 1
        
        measures = plan.measures
        
        def write_measure(code, label, item):
            nonlocal row
            ws.cell(row=row, column=1, value=f"{code}. {label}").font = Font(bold=True); row += 1
            write_row(row, "現在の取組", item.current_measure, "Measures"); row += 1
            write_row(row, "今後の計画", item.future_plan, "Measures"); row += 1

        write_measure("A", "自然災害等が発生した場合における人員体制の整備", measures.personnel)
        write_measure("B", "事業継続力強化に資する設備、機器及び装置の導入", measures.building)
        write_measure("C", "事業活動を継続するための資金の調達手段の確保", measures.money)
        write_measure("D", "事業活動を継続するための重要情報の保護", measures.data)
        row += 1
        
        # Section 6: Finance
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【様式第6】 資金計画").font = Font(bold=True)
        row += 1
        
        for i, f in enumerate(plan.financial_plan.items, 1):
            write_row(row, f"{i}. 実施事項", f.item, "FinancialPlan"); row += 1
            write_row(row, f"   使途・用途", f.usage, "FinancialPlan"); row += 1
            write_row(row, f"   資金調達方法", f.method, "FinancialPlan"); row += 1
            write_row(row, f"   金額 (円)", f.amount, "FinancialPlan"); row += 1
        row += 1
        
        # PDCA
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【推進体制】").font = Font(bold=True)
        row += 1
        
        write_row(row, "経営層の下推進 (平時の推進体制)", plan.pdca.management_system, "PDCA"); row += 1
        write_row(row, "教育・訓練", plan.pdca.training_education, "PDCA"); row += 1
        write_row(row, "見直しを計画", plan.pdca.plan_review, "PDCA"); row += 1
        
        # Period & Applicant Info
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="【その他】 申請情報").font = Font(bold=True)
        row += 1
        
        write_row(row, "実施期間(開始)", plan.period.start_date, "BasicInfo"); row += 1
        write_row(row, "実施期間(終了)", plan.period.end_date, "BasicInfo"); row += 1
        write_row(row, "担当者名", plan.applicant_info.contact_name, "BasicInfo"); row += 1
        write_row(row, "メールアドレス", plan.applicant_info.email, "BasicInfo"); row += 1
        write_row(row, "電話番号", plan.applicant_info.phone, "BasicInfo"); row += 1
        write_row(row, "決算月", plan.applicant_info.closing_month, "BasicInfo"); row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_for_application(plan: ApplicationRoot) -> BytesIO:
        """
        Create Excel sheet optimized for direct copy-paste into electronic application.
        Format: Field Name | Value (copy this)
        
        Args:
            plan: ApplicationRoot instance
            
        Returns:
            BytesIO object containing Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "電子申請入力用"
        
        # Styles - minimal, focused on readability
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        section_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
        value_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Yellow for copy
        header_font = Font(color="FFFFFF", bold=True, size=12)
        section_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def write_field(row_num, field_name, value):
            """Write a field with copy-friendly formatting"""
            cell_label = ws.cell(row=row_num, column=1, value=field_name)
            cell_label.border = border
            cell_label.alignment = Alignment(vertical='center')
            
            # Value cell - highlighted for easy copy
            cell_value = ws.cell(row=row_num, column=2, value=value if value else "")
            cell_value.fill = value_fill
            cell_value.border = border
            cell_value.alignment = Alignment(vertical='center', wrap_text=True)
            return row_num + 1
        
        def write_section(row_num, title):
            """Write section header"""
            ws.merge_cells(f'A{row_num}:B{row_num}')
            cell = ws.cell(row=row_num, column=1, value=title)
            cell.fill = section_fill
            cell.font = section_font
            cell.border = border
            return row_num + 1
        
        row = 1
        
        # Title
        ws.merge_cells('A1:B1')
        title_cell = ws.cell(row=1, column=1, value="📋 電子申請 入力用シート（コピペ用）")
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal='center')
        row = 3
        
        # Instructions
        ws.merge_cells('A2:B2')
        ws.cell(row=2, column=1, value="黄色のセルの値をコピーして電子申請システムに貼り付けてください").font = Font(italic=True, color="666666")
        
        # === Section 1: 基本情報 ===
        row = write_section(row, "【様式第1】基本情報")
        bi = plan.basic_info
        row = write_field(row, "事業者の氏名又は名称", bi.corporate_name)
        row = write_field(row, "フリガナ", bi.corporate_name_kana)
        row = write_field(row, "法人番号（13桁）", bi.corporate_number)
        row = write_field(row, "設立年月日", bi.establishment_date)
        row = write_field(row, "代表者の役職", bi.representative_title)
        row = write_field(row, "代表者の氏名", bi.representative_name)
        row = write_field(row, "郵便番号", bi.address_zip)
        row = write_field(row, "都道府県", bi.address_pref)
        row = write_field(row, "市区町村", bi.address_city)
        row = write_field(row, "住所（字・番地等）", bi.address_street)
        row = write_field(row, "マンション名等", bi.address_building or "")
        row = write_field(row, "業種（大分類）", bi.industry_major)
        row = write_field(row, "業種（中分類）", bi.industry_middle)
        row = write_field(row, "資本金又は出資の額（円）", f"{bi.capital:,}" if bi.capital else "")
        row = write_field(row, "常時使用する従業員の数", bi.employees)
        row += 1
        
        # === Section 2: 事業概要・目標 ===
        row = write_section(row, "【様式第2】事業活動の概要・取組目的")
        row = write_field(row, "自社の事業活動の概要", plan.goals.business_overview)
        row = write_field(row, "事業継続力強化に取り組む目的", plan.goals.business_purpose)
        row += 1
        
        # === Section 3: 災害想定 ===
        row = write_section(row, "【様式第3】自然災害等の想定")
        ds = plan.goals.disaster_scenario
        row = write_field(row, "事業活動に影響を与える自然災害等の想定", ds.disaster_assumption)
        row = write_field(row, "影響（人員）", ds.impacts.impact_personnel if ds.impacts else "")
        row = write_field(row, "影響（建物・設備）", ds.impacts.impact_building if ds.impacts else "")
        row = write_field(row, "影響（資金繰り）", ds.impacts.impact_funds if ds.impacts else "")
        row = write_field(row, "影響（情報）", ds.impacts.impact_info if ds.impacts else "")
        row += 1
        
        # === Section 4: 初動対応 ===
        row = write_section(row, "【様式第4】初動対応の取組内容")
        
        # Fixed 4 categories for electronic application
        categories = [
            ("人命の安全確保（避難）", "action_content", "timing", "preparation_content"),
            ("人命の安全確保（安否確認）", "action_content", "timing", "preparation_content"),
            ("非常時の緊急時体制の構築", "action_content", "timing", "preparation_content"),
            ("被害状況の把握・共有", "action_content", "timing", "preparation_content"),
        ]
        
        # Try to match response_procedures to categories
        rp_list = plan.response_procedures
        
        def find_rp(keywords):
            for p in rp_list:
                cat = getattr(p, 'category', '') or ''
                content = getattr(p, 'action_content', '') or ''
                if any(k in cat or k in content for k in keywords):
                    return p
            return None
        
        # 1. 人命の安全確保（避難）
        evacuation = find_rp(["避難", "人命", "安全確保"])
        row = write_field(row, "1. 人命の安全確保（避難）- 取組内容", getattr(evacuation, 'action_content', '') if evacuation else "")
        row = write_field(row, "   発災後の対応時期", getattr(evacuation, 'timing', '') if evacuation else "")
        row = write_field(row, "   事前対策", getattr(evacuation, 'preparation_content', '') if evacuation else "")
        
        # 2. 人命の安全確保（安否確認）
        confirmation = find_rp(["安否", "確認"])
        row = write_field(row, "2. 人命の安全確保（安否確認）- 取組内容", getattr(confirmation, 'action_content', '') if confirmation else "")
        row = write_field(row, "   発災後の対応時期", getattr(confirmation, 'timing', '') if confirmation else "")
        row = write_field(row, "   事前対策", getattr(confirmation, 'preparation_content', '') if confirmation else "")
        
        # 3. 緊急時体制
        org = find_rp(["体制", "本部", "緊急"])
        row = write_field(row, "3. 非常時の緊急時体制 - 取組内容", getattr(org, 'action_content', '') if org else "")
        row = write_field(row, "   発災後の対応時期", getattr(org, 'timing', '') if org else "")
        row = write_field(row, "   事前対策", getattr(org, 'preparation_content', '') if org else "")
        
        # 4. 被害状況把握
        info = find_rp(["被害", "把握", "情報", "連絡"])
        row = write_field(row, "4. 被害状況の把握・共有 - 取組内容", getattr(info, 'action_content', '') if info else "")
        row = write_field(row, "   発災後の対応時期", getattr(info, 'timing', '') if info else "")
        row = write_field(row, "   事前対策", getattr(info, 'preparation_content', '') if info else "")
        row += 1
        
        # === Section 5: 事前対策 ===
        row = write_section(row, "【様式第5】事前対策の取組内容")
        m = plan.measures
        row = write_field(row, "A. 人員体制の整備 - 現在の取組", m.personnel.current_measure if m.personnel else "")
        row = write_field(row, "A. 人員体制の整備 - 今後の計画", m.personnel.future_plan if m.personnel else "")
        row = write_field(row, "B. 設備等の導入 - 現在の取組", m.building.current_measure if m.building else "")
        row = write_field(row, "B. 設備等の導入 - 今後の計画", m.building.future_plan if m.building else "")
        row = write_field(row, "C. 資金の調達手段確保 - 現在の取組", m.money.current_measure if m.money else "")
        row = write_field(row, "C. 資金の調達手段確保 - 今後の計画", m.money.future_plan if m.money else "")
        row = write_field(row, "D. 重要情報の保護 - 現在の取組", m.data.current_measure if m.data else "")
        row = write_field(row, "D. 重要情報の保護 - 今後の計画", m.data.future_plan if m.data else "")
        row += 1
        
        # === Section 6: 推進体制 ===
        row = write_section(row, "【様式第6】平時の推進体制")
        pdca = plan.pdca
        row = write_field(row, "経営層関与の下での推進体制", pdca.management_system)
        row = write_field(row, "教育及び訓練の実施", pdca.training_education)
        row = write_field(row, "教育・訓練の実施月", f"{pdca.training_month}月" if pdca.training_month else "")
        row = write_field(row, "計画の見直し", pdca.plan_review)
        row = write_field(row, "計画見直しの実施月", f"{pdca.review_month}月" if pdca.review_month else "")
        row = write_field(row, "取組の社内周知", pdca.internal_publicity)
        row += 1
        
        # === Section 7: 資金計画 ===
        row = write_section(row, "【様式第7】資金計画")
        for i, fp in enumerate(plan.financial_plan.items, 1):
            row = write_field(row, f"{i}. 実施事項", fp.item)
            row = write_field(row, f"   使途・用途", fp.usage)
            row = write_field(row, f"   資金調達方法", fp.method)
            row = write_field(row, f"   金額（円）", f"{fp.amount:,}" if fp.amount else "")
        if not plan.financial_plan.items:
            row = write_field(row, "資金計画", "（未入力）")
        row += 1
        
        # === Section 8: 実施期間・担当者 ===
        row = write_section(row, "【その他】実施期間・担当者情報")
        row = write_field(row, "実施期間（開始）", plan.period.start_date)
        row = write_field(row, "実施期間（終了）", plan.period.end_date)
        row = write_field(row, "担当者名", plan.applicant_info.contact_name)
        row = write_field(row, "メールアドレス", plan.applicant_info.email)
        row = write_field(row, "電話番号", plan.applicant_info.phone)
        
        # Column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 70
        
        # Freeze panes for easier navigation
        ws.freeze_panes = 'A3'
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

