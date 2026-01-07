"""
E2E Test Script: Chat History to Draft Sheet Export
Uses the session_default.json chat history to generate an Excel draft sheet
"""
import json
import os
import sys

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from datetime import datetime

def load_chat_history():
    """Load the chat history from session_default.json"""
    session_path = os.path.join(os.path.dirname(__file__), "userdata", "session_default.json")
    with open(session_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data

def extract_structured_data(history):
    """Extract the structured data from the last model message"""
    for msg in reversed(history):
        if msg.get("role") == "model" and "<data>" in msg.get("content", ""):
            content = msg["content"]
            start = content.find("<data>") + len("<data>")
            end = content.find("</data>")
            if start > 0 and end > 0:
                json_str = content[start:end].strip()
                return json.loads(json_str)
    return None

def generate_excel(data, output_path):
    """Generate Excel draft sheet from structured data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Sheet"
    
    # Title
    ws["A1"] = "Event Continuity Strength Plan - Draft Sheet"
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Basic Info Section
    row = 4
    ws[f"A{row}"] = "1. Basic Information"
    row += 1
    
    basic = data.get("basic_info", {})
    fields = [
        ("Company Name", basic.get("company_name", "")),
        ("Company Name (Kana)", basic.get("company_name_furigana", "")),
        ("Corporate Number", basic.get("corporate_number", "")),
        ("Establishment Date", basic.get("establishment_date", "")),
        ("Postal Code", basic.get("postal_code", "")),
        ("Address", basic.get("address", "")),
        ("Representative Position", basic.get("representative_position", "")),
        ("Representative Name", basic.get("representative_name", "")),
        ("Industry (Major)", basic.get("industry_major_category", "")),
        ("Industry (Middle)", basic.get("industry_middle_category", "")),
        ("Capital Amount", basic.get("capital_amount", "")),
        ("Employee Count", basic.get("employee_count", "")),
    ]
    
    for label, value in fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Business Overview Section
    row += 1
    ws[f"A{row}"] = "2. Business Overview"
    row += 1
    
    overview = data.get("business_overview", {})
    ws[f"A{row}"] = "Activity Summary"
    ws[f"B{row}"] = overview.get("activity_summary", "")
    row += 1
    ws[f"A{row}"] = "Purpose of Plan"
    ws[f"B{row}"] = overview.get("purpose_of_plan", "")
    row += 1
    
    # Disaster Scenario Section
    row += 1
    ws[f"A{row}"] = "3. Disaster Scenario"
    row += 1
    
    disaster = data.get("disaster_scenario", {})
    disaster_fields = [
        ("Disaster Type", disaster.get("disaster_type", "")),
        ("Human Impact", disaster.get("damage_human", "")),
        ("Material Impact", disaster.get("damage_material", "")),
        ("Financial Impact", disaster.get("damage_money", "")),
        ("Information Impact", disaster.get("damage_information", "")),
    ]
    
    for label, value in disaster_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Initial Response Section
    row += 1
    ws[f"A{row}"] = "4. Initial Response"
    row += 1
    
    initial = data.get("initial_response", {})
    init_fields = [
        ("Human Safety", initial.get("safety_human", "")),
        ("Emergency System", initial.get("emergency_system", "")),
        ("Damage Assessment", initial.get("damage_assessment", "")),
    ]
    
    for label, value in init_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Measures Section
    row += 1
    ws[f"A{row}"] = "5. Measures"
    row += 1
    
    measures = data.get("measures", {})
    measure_fields = [
        ("Human Measures", measures.get("human_measures", "")),
        ("Material Measures", measures.get("material_measures", "")),
        ("Financial Measures", measures.get("money_measures", "")),
        ("Information Measures", measures.get("information_measures", "")),
    ]
    
    for label, value in measure_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Implementation System Section
    row += 1
    ws[f"A{row}"] = "6. Implementation System"
    row += 1
    
    impl = data.get("implementation_system", {})
    impl_fields = [
        ("Normal Time System", impl.get("normal_time_system", "")),
        ("Training Month", impl.get("training_month", "")),
        ("Review Month", impl.get("review_month", "")),
    ]
    
    for label, value in impl_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Finance and Period Section
    row += 1
    ws[f"A{row}"] = "7. Finance and Period"
    row += 1
    
    finance = data.get("finance_and_period", {})
    finance_fields = [
        ("Implementation Period", finance.get("implementation_period", "")),
        ("Estimated Cost", finance.get("estimated_cost", "")),
        ("Funding Method", finance.get("funding_method", "")),
    ]
    
    for label, value in finance_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Contact Info Section
    row += 1
    ws[f"A{row}"] = "8. Contact Information"
    row += 1
    
    contact = data.get("contact_info", {})
    contact_fields = [
        ("Person in Charge", contact.get("person_in_charge_name", "")),
        ("Furigana", contact.get("person_in_charge_furigana", "")),
        ("Email", contact.get("email_address", "")),
        ("Phone", contact.get("phone_number", "")),
    ]
    
    for label, value in contact_fields:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = str(value)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    
    # Save
    wb.save(output_path)
    print(f"[OK] Draft sheet saved to: {output_path}")
    return output_path

def main():
    print("=" * 60)
    print("E2E Test: Chat History to Draft Sheet Export")
    print("=" * 60)
    
    # Load chat history
    print("[*] Loading chat history from session_default.json...")
    session_data = load_chat_history()
    history = session_data.get("history", [])
    print(f"[OK] Loaded {len(history)} messages")
    
    # Extract structured data
    print("[*] Extracting structured data from chat...")
    structured_data = extract_structured_data(history)
    
    if not structured_data:
        print("[FAIL] Could not extract structured data from chat history")
        return
    
    print(f"[OK] Extracted data for: {structured_data.get('basic_info', {}).get('company_name', 'Unknown')}")
    
    # Generate Excel
    output_path = os.path.join(os.path.dirname(__file__), "output_draft_sheet.xlsx")
    print(f"[*] Generating Excel draft sheet...")
    generate_excel(structured_data, output_path)
    
    print("=" * 60)
    print("[SUCCESS] E2E Test Complete!")
    print(f"Output file: {output_path}")

if __name__ == "__main__":
    main()
